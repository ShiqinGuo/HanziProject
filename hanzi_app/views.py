from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.core.paginator import Paginator
from django.http import JsonResponse, FileResponse, HttpResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.conf import settings
from django.db import transaction
from django.core.cache import cache
from django.db.models import Q, Count, Prefetch
from django.db.models import Value, F  
from django.db.models import IntegerField  
import json
import os
import shutil
import zipfile
from .models import Hanzi
from .forms import HanziForm
import time
from django.views.decorators.cache import cache_page
from .generate import generate_hanzi_image, get_stroke_order,get_pinyin
import pandas as pd
from io import BytesIO
import logging
import traceback
import random
import csv
import tempfile
from datetime import datetime
from rest_framework.serializers import ModelSerializer
from django.contrib import messages
from rest_framework import serializers
from urllib.parse import urlencode
import re
import openpyxl
import uuid
from django.core.files.storage import FileSystemStorage
from hanzi_app.data_importer import import_hanzi_data, clean_import_results_folder, extract_zip_to_temp
from concurrent.futures import ThreadPoolExecutor
import glob
from celery.result import AsyncResult
from celery.app.control import Inspect
from hanzi_project.celery import app as celery_app
import redis
import platform
import celery.worker.control

# 在文件顶部添加日志配置
logger = logging.getLogger(__name__)

# 为导出功能添加序列化器
class HanziSerializer(ModelSerializer):
    # 自定义字段，处理image_path格式
    image_path = serializers.SerializerMethodField()
    
    class Meta:
        model = Hanzi
        fields = ['character', 'structure', 'variant', 'level', 'comment', 'image_path']
    
    def get_image_path(self, obj):
        """处理图片路径，只返回文件名部分（不含后缀）"""
        if not obj.image_path:
            return ""
        filename = os.path.basename(obj.image_path)
        return os.path.splitext(filename)[0]  # 不含后缀的文件名

# 预加载笔画数据
stroke_dict = {}
stoke_file_path = os.path.join(settings.BASE_DIR, 'data\\ch_match\\stoke.txt')
try:
    with open(stoke_file_path, 'r', encoding='utf-8') as f:
        for line in f:
            parts = line.strip().split('|')
            stroke_dict[parts[1]] = parts[2]
except FileNotFoundError:
    print(f"文件未找到: {stoke_file_path}")

# 定义上传文件夹路径
UPLOAD_FOLDER = os.path.join(settings.MEDIA_ROOT, 'uploads')  # 直接使用media根目录
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 定义前端日志目录
FRONTEND_LOGS_FOLDER = os.path.join(settings.BASE_DIR, 'logs', 'frontend')
if not os.path.exists(FRONTEND_LOGS_FOLDER):
    os.makedirs(FRONTEND_LOGS_FOLDER, exist_ok=True)

# 添加前端日志捕获视图
@csrf_exempt
def capture_frontend_logs(request):
    """捕获前端控制台日志并保存到服务器"""
    if request.method == 'POST':
        try:
            log_data = json.loads(request.body)
            log_level = log_data.get('level', 'info')  # 日志级别：info, warn, error等
            log_message = log_data.get('message', '')
            page_url = log_data.get('url', '')
            timestamp = log_data.get('timestamp', datetime.now().isoformat())
            
            # 创建对应日期的日志文件
            today = datetime.now().strftime('%Y-%m-%d')
            log_file_path = os.path.join(FRONTEND_LOGS_FOLDER, f'frontend-logs-{today}.log')
            
            # 格式化日志消息
            formatted_log = f"[{timestamp}] [{log_level.upper()}] [{page_url}] {log_message}\n"
            
            # 写入日志文件
            with open(log_file_path, 'a', encoding='utf-8') as f:
                f.write(formatted_log)
            
            # 记录到Django日志系统
            if log_level == 'error':
                logger.error(f"前端错误: {log_message}")
            elif log_level == 'warn':
                logger.warning(f"前端警告: {log_message}")
            else:
                logger.info(f"前端日志: {log_message}")
                
            return JsonResponse({'success': True, 'message': '日志已记录'})
            
        except Exception as e:
            logger.error(f"捕获前端日志失败: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'error': '仅支持POST请求'}, status=405)

@csrf_exempt
@login_required
def delete_logs(request):
    """删除前端日志文件"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST请求'}, status=405)
    
    try:
        # 清空所有日志
        if request.POST.get('all') == 'true':
            log_files = [f for f in os.listdir(FRONTEND_LOGS_FOLDER) 
                         if f.startswith('frontend-logs-') and f.endswith('.log')]
            deleted_count = 0
            
            for file in log_files:
                file_path = os.path.join(FRONTEND_LOGS_FOLDER, file)
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_count += 1
            
            return JsonResponse({
                'success': True, 
                'message': f'已清空所有{deleted_count}个日志文件',
                'count': deleted_count
            })
        
        # 删除指定日期的日志
        date = request.POST.get('date')
        level = request.POST.get('level', 'all')
        
        if date:
            log_file_path = os.path.join(FRONTEND_LOGS_FOLDER, f'frontend-logs-{date}.log')
            
            # 如果是按级别删除，则需要读取日志内容，过滤掉指定级别的日志，然后重写文件
            if level and level != 'all' and os.path.exists(log_file_path):
                with open(log_file_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                
                # 按级别过滤日志
                filtered_lines = []
                deleted_count = 0
                
                for line in lines:
                    match = re.match(r'\[(.*?)\] \[(.*?)\] \[(.*?)\] (.*)', line.strip())
                    if match:
                        log_level = match.group(2).lower()
                        if log_level != level.lower():
                            filtered_lines.append(line)
                        else:
                            deleted_count += 1
                    else:
                        filtered_lines.append(line)
                
                # 如果有删除操作，重写文件
                if deleted_count > 0:
                    with open(log_file_path, 'w', encoding='utf-8') as f:
                        f.writelines(filtered_lines)
                    
                    return JsonResponse({
                        'success': True, 
                        'message': f'已删除{date}日志中的{deleted_count}条{level.upper()}级别日志',
                        'count': deleted_count
                    })
                else:
                    return JsonResponse({
                        'success': False, 
                        'message': f'未找到{date}日志中的{level.upper()}级别日志'
                    })
            
            # 如果删除整个日期的日志文件
            elif os.path.exists(log_file_path):
                os.remove(log_file_path)
                return JsonResponse({
                    'success': True, 
                    'message': f'已删除{date}的日志文件',
                    'count': 1
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': f'未找到{date}的日志文件'
                })
        
        return JsonResponse({'success': False, 'message': '缺少必要的参数'})
    
    except Exception as e:
        logger.error(f"删除日志失败: {str(e)}")
        return JsonResponse({'success': False, 'message': f'删除日志时发生错误: {str(e)}'}, status=500)

@login_required
def view_frontend_logs(request):
    """查看前端日志文件列表及内容"""
    date = request.GET.get('date')
    level = request.GET.get('level', 'all')
    search = request.GET.get('search', '')
    
    log_files = []
    logs = []
    
    # 获取所有日志文件
    for file in os.listdir(FRONTEND_LOGS_FOLDER):
        if file.startswith('frontend-logs-') and file.endswith('.log'):
            log_date = file[14:-4]  # 提取日期部分
            log_files.append({
                'date': log_date,
                'file': file,
                'size': os.path.getsize(os.path.join(FRONTEND_LOGS_FOLDER, file)) // 1024,  # KB
                'last_modified': datetime.fromtimestamp(os.path.getmtime(os.path.join(FRONTEND_LOGS_FOLDER, file)))
            })
    
    # 按日期降序排序
    log_files.sort(key=lambda x: x['date'], reverse=True)
    
    # 如果没有指定日期，使用最新的日志文件
    if not date and log_files:
        date = log_files[0]['date']
    
    # 如果指定了日期，则显示该日期的日志内容
    if date:
        log_file_path = os.path.join(FRONTEND_LOGS_FOLDER, f'frontend-logs-{date}.log')
        if os.path.exists(log_file_path):
            try:
                with open(log_file_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    
                # 解析和过滤日志
                for line in lines:
                    try:
                        # 解析日志行
                        match = re.match(r'\[(.*?)\] \[(.*?)\] \[(.*?)\] (.*)', line.strip())
                        if match:
                            timestamp, log_level, url, message = match.groups()
                            log_level = log_level.lower()
                            
                            # 应用过滤器
                            if (level == 'all' or log_level.lower() == level) and \
                               (not search or search.lower() in message.lower() or search.lower() in url.lower()):
                                logs.append({
                                    'timestamp': timestamp,
                                    'level': log_level.lower(),
                                    'url': url,
                                    'message': message
                                })
                        else:
                            # 如果不匹配格式，将整行作为消息处理
                            if not search or search.lower() in line.lower():
                                logs.append({
                                    'timestamp': '',
                                    'level': 'info',
                                    'url': '',
                                    'message': line.strip()
                                })
                    except Exception as e:
                        logger.error(f"解析日志行失败: {str(e)}, 行内容: {line}")
                        logs.append({
                            'timestamp': '',
                            'level': 'error',
                            'url': '',
                            'message': f"解析错误: {line.strip()}"
                        })
            except Exception as e:
                logger.error(f"读取日志文件失败: {str(e)}")
                messages.error(request, f"读取日志文件失败: {str(e)}")
    
    context = {
        'log_files': log_files,
        'logs': logs,
        'selected_date': date,
        'selected_level': level,
        'search_query': search
    }
    
    return render(request, 'hanzi_app/view_logs.html', context)

@login_required
def index(request):
    # 获取筛选参数
    search = request.GET.get('search', '')
    structure = request.GET.get('structure', '')
    level = request.GET.get('level', '')
    variant = request.GET.get('variant', '')
    stroke_count = request.GET.get('stroke_count', '')
    page_number = request.GET.get('page', 1)
    
    # 如果是从详情页返回，且没有筛选参数，尝试从会话中恢复
    is_returning = request.GET.get('returning', '0') == '1'
    if is_returning and not any([search, structure, level, variant, stroke_count]) and 'last_filter' in request.session:
        last_filter = request.session['last_filter']
        search = last_filter.get('search', '')
        structure = last_filter.get('structure', '')
        level = last_filter.get('level', '')
        variant = last_filter.get('variant', '')
        stroke_count = last_filter.get('stroke_count', '')
        page_number = last_filter.get('page', 1)
    
    # 保存当前筛选条件到会话
    request.session['last_filter'] = {
        'search': search,
        'structure': structure,
        'level': level,
        'variant': variant,
        'stroke_count': stroke_count,
        'page': page_number
    }
    
    # 移除缓存相关代码，直接查询数据库
    hanzi_list = Hanzi.objects.all().order_by('id')
    
    # 应用筛选条件
    if search:
        hanzi_list = hanzi_list.filter(
            Q(character__contains=search) | 
            Q(pinyin__contains=search) |
            Q(id__contains=search)
        )
    
    if structure and structure != '所有':
        hanzi_list = hanzi_list.filter(structure=structure)
    
    if level and level != '所有':
        hanzi_list = hanzi_list.filter(level=level)
        
    if variant and variant != '所有':
        hanzi_list = hanzi_list.filter(variant=variant)
    
    if stroke_count and stroke_count != '所有':
        hanzi_list = hanzi_list.filter(stroke_count=int(stroke_count))
    
    # 为每个汉字添加动画延迟
    for i, hanzi in enumerate(hanzi_list):
        page_i = i % 20
        hanzi.animation_delay = page_i * 50  
    
    # 分页处理
    paginator = Paginator(hanzi_list, 20)  # 每页显示20条
    page_obj = paginator.get_page(page_number)
    
    # 准备上下文数据
    context = {
        'page_obj': page_obj,
        'search': search,
        'structure': structure,
        'level': level,
        'variant': variant,
        'selected_structure': structure,
        'selected_level': level,
        'selected_variant': variant,
        'selected_stroke': stroke_count,
        'structure_choices': Hanzi.STRUCTURE_CHOICES,
        'level_choices': Hanzi.LEVEL_CHOICES,
        'variant_choices': Hanzi.VARIANT_CHOICES,
        'stroke_count_options': list(range(1, 21)),  # 1-20的笔画范围
        'total_count': hanzi_list.count(),
        'hanzi_list': hanzi_list,
    }
    
    return render(request, 'hanzi_app/index.html', context)

@cache_page(60 * 15)  # 缓存15分钟
def get_stroke_count(request, char):
    # 使用缓存键
    cache_key = f'stroke_count_{char}'
    stroke_count = cache.get(cache_key)
    
    if stroke_count is None:
        # 缓存未命中，从数据库或文件获取
        stroke_count = stroke_dict.get(char, '0')
        # 存入缓存，有效期1天
        cache.set(cache_key, stroke_count, 60*60*24)
    
    return JsonResponse({'stroke_count': stroke_count})

@csrf_exempt
@require_http_methods(['POST'])
def generate_id(request):
    structure_map = {
        "未知结构": "0",
        "左右结构": "1",  # 结构类型映射前缀
        "上下结构": "2",
        "包围结构": "3",
        "独体结构": "4",
        "品字结构": "5",
        "穿插结构": "6"
    }
    
    try:
        # 从前端请求的JSON body中获取结构类型
        data = json.loads(request.body)
        structure = data.get('structure')
        # 验证结构数据有效性
        if structure not in structure_map:
            return JsonResponse({"error": "无效结构类型"}, status=400)
        # 获取该结构类型的最新ID
        prefix = structure_map[structure]
        last_entry = Hanzi.objects.filter(id__startswith=prefix).order_by('-id').first()
        
        # 核心逻辑：自增编号
        if last_entry:
            # 示例：最后ID是"10015" → 提取"0015" → 转换为15 → +1=16
            last_num = int(last_entry.id[1:]) 
            new_num = last_num + 1
        else:
            # 该结构首次使用时从1开始
            new_num = 1
        
        # 生成新ID（前缀+4位数字）
        new_id = f"{prefix}{new_num:04d}"
        return JsonResponse({"id": new_id})
    
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif'}

def generate_filename(generated_id, filename):
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'unknown'
    return f"{generated_id}.{ext}"

def remove_existing_files(filepath):
    try:
        full_path = os.path.join(UPLOAD_FOLDER, filepath)  # 直接使用MEDIA_ROOT
        if os.path.exists(full_path):
            os.remove(full_path)
    except Exception as e:
        print(f"删除文件失败: {e}")

@csrf_exempt
def add_hanzi(request):
    if request.method == 'POST':
        try:
            character = request.POST.get('character')
            generated_id = request.POST.get('generated_id')
            
            # 验证汉字字符
            if not character or len(character) != 1:
                return JsonResponse({'error': '请输入单个汉字字符'}, status=400)
            
            # 验证ID
            if not generated_id:
                return JsonResponse({'error': '请生成有效的ID'}, status=400)
            
            # 处理用户上传的图片
            image_file = request.FILES.get('image_file')
            if not image_file:
                return JsonResponse({'error': '请上传用户书写的汉字图片'}, status=400)
            
            # 验证图片格式
            if not allowed_file(image_file.name):
                return JsonResponse({'error': '不支持的图片格式'}, status=400)
            
            # 保存用户图片
            filename = generate_filename(generated_id, image_file.name)
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            
            with open(file_path, 'wb+') as destination:
                for chunk in image_file.chunks():
                    destination.write(chunk)
            
            # 自动获取笔顺
            stroke_order = request.POST.get('stroke_order', '')
            if not stroke_order:
                # 如果用户没有输入笔顺，尝试自动获取
                stroke_orders = get_stroke_order(character)
                if stroke_orders and stroke_orders[0]:
                    stroke_order = stroke_orders[0]
            
            # 创建汉字对象
            hanzi = Hanzi(
                id=generated_id,
                character=character,
                stroke_count=int(request.POST.get('stroke_count', 0)),
                structure=request.POST.get('structure', '未知结构'),
                pinyin=request.POST.get('pinyin', ''),
                level=request.POST.get('level', 'D'),
                variant=request.POST.get('variant', '简体'),
                comment=request.POST.get('comment', ''),
                image_path=f"uploads/{filename}",
                stroke_order=stroke_order
            )
            hanzi.save()
            
            # 自动生成标准图片并更新数据库
            standard_path = generate_hanzi_image(hanzi.character)
            if standard_path:
                # 提取相对路径并更新数据库
                rel_path = os.path.join("standard_images", f"{hanzi.character}.jpg")
                Hanzi.objects.filter(id=hanzi.id).update(standard_image=rel_path)
            
            return redirect('hanzi_app:index')
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return render(request, 'hanzi_app/add.html')

def hanzi_detail(request, hanzi_id):
    hanzi = get_object_or_404(Hanzi, id=hanzi_id)
    
    # 构建返回URL，添加返回标记并保留所有筛选条件
    url_params = {
        'returning': '1',
    }
    
    # 保留所有筛选参数和分页参数
    filter_params = ['search', 'structure', 'level', 'variant', 'stroke_count', 'page']
    for param in filter_params:
        value = request.GET.get(param)
        if value:
            url_params[param] = value
    
    # 构建完整的返回URL
    back_url = reverse('hanzi_app:index') + '?' + urlencode(url_params)
    
    # 获取汉字的笔画顺序
    if hanzi.stroke_order:
        stroke_order = hanzi.stroke_order
    else:
        stroke_order = get_auto_stroke_order(hanzi.character)
        
    # 获取拼音
    pinyin = hanzi.pinyin if hanzi.pinyin else get_pinyin(hanzi.character)
    
    context = {
        'hanzi': hanzi,
        'stroke_order': stroke_order,
        'pinyin': pinyin,
        'back_url': back_url
    }
    return render(request, 'hanzi_app/detail.html', context)

def delete_hanzi(request, hanzi_id):
    try:
        hanzi = get_object_or_404(Hanzi, id=hanzi_id)
        
        # 删除对应的图片文件
        if hanzi.image_path:
            try:
                file_path = os.path.join(UPLOAD_FOLDER, os.path.basename(hanzi.image_path))
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                logger.error(f"删除文件失败: {e}")
        
        # 删除数据库记录
        hanzi.delete()
        
        # 构建返回URL，添加返回标记并保留所有筛选条件
        url_params = {
            'returning': '1',
        }
        
        # 保留所有筛选参数和分页参数
        filter_params = ['search', 'structure', 'level', 'variant', 'stroke_count', 'page']
        for param in filter_params:
            value = request.GET.get(param)
            if value:
                url_params[param] = value
        
        # 构建完整的返回URL并重定向
        redirect_url = reverse('hanzi_app:index') + '?' + urlencode(url_params)
        return redirect(redirect_url)
    except Exception as e:
        # 记录异常
        logger.error(f"删除汉字失败: {e}")
        return HttpResponse(f"删除失败: {str(e)}", status=500)

def edit_hanzi(request, hanzi_id):
    hanzi = get_object_or_404(Hanzi, id=hanzi_id)
    
    # 添加模板需要的选项
    structure_options = [choice[0] for choice in Hanzi.STRUCTURE_CHOICES]
    variant_options = [choice[0] for choice in Hanzi.VARIANT_CHOICES]
    
    # 构建返回URL，添加返回标记并保留所有筛选条件
    url_params = {
        'returning': '1',
    }
    
    # 保留所有筛选参数和分页参数
    filter_params = ['search', 'structure', 'level', 'variant', 'stroke_count', 'page']
    for param in filter_params:
        value = request.GET.get(param)
        if value:
            url_params[param] = value
    
    # 构建完整的返回URL
    back_url = reverse('hanzi_app:index') + '?' + urlencode(url_params)
    
    if request.method == 'POST' and not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        form = HanziForm(request.POST, request.FILES, instance=hanzi)
        if form.is_valid():
            try:
                # 开启事务处理
                with transaction.atomic():
                    hanzi_instance = form.save(commit=False)
                    
                    # 检查结构是否发生变化
                    original_structure = hanzi.structure
                    new_structure = hanzi_instance.structure
                    if new_structure != original_structure:
                        # 结构变化，生成新ID
                        old_id = hanzi_instance.id
                        new_id = generate_new_id(new_structure)
                        logger.info(f"汉字'{hanzi_instance.character}'结构从'{original_structure}'变为'{new_structure}'，ID从'{old_id}'更新为'{new_id}'")
                        
                        # 更新ID
                        hanzi_instance.id = new_id
                    
                    # 处理图片上传 - 用户图片
                    if 'new_image_file' in request.FILES:
                        # 先删除旧图片
                        if hanzi.image_path:
                            old_image_path = os.path.join(UPLOAD_FOLDER, os.path.basename(hanzi.image_path))
                            if os.path.exists(old_image_path):
                                os.remove(old_image_path)
                        
                        image_file = request.FILES['new_image_file']
                        filename = f"{hanzi_instance.id}.{image_file.name.split('.')[-1]}"
                        file_path = os.path.join(UPLOAD_FOLDER, filename)
                        
                        with open(file_path, 'wb+') as destination:
                            for chunk in image_file.chunks():
                                destination.write(chunk)
                        
                        # 更新图片路径
                        hanzi_instance.image_path = f"uploads/{filename}"
                    elif new_structure != original_structure and hanzi.image_path:
                        # 结构变化但没有上传新图片，需要更新现有图片文件名
                        old_image_path = os.path.join(UPLOAD_FOLDER, os.path.basename(hanzi.image_path))
                        if os.path.exists(old_image_path):
                            # 获取文件扩展名
                            filename_parts = os.path.basename(hanzi.image_path).split('.')
                            if len(filename_parts) > 1:
                                ext = filename_parts[-1]
                                # 新文件名使用新ID
                                new_filename = f"{hanzi_instance.id}.{ext}"
                                new_image_path = os.path.join(UPLOAD_FOLDER, new_filename)
                                # 复制文件
                                shutil.copy2(old_image_path, new_image_path)
                                # 删除旧文件
                                os.remove(old_image_path)
                                logger.info(f"已重命名图片文件: {os.path.basename(old_image_path)} -> {new_filename}")
                                # 更新数据库中的图片路径
                                hanzi_instance.image_path = f"uploads/{new_filename}"
                    
                    # 处理标准图片上传
                    if 'new_standard_file' in request.FILES:
                        # 标准图片处理逻辑
                        standard_file = request.FILES['new_standard_file']
                        standard_filename = f"{hanzi_instance.id}_standard.{standard_file.name.split('.')[-1]}"
                        standard_file_path = os.path.join(UPLOAD_FOLDER, standard_filename)
                        
                        with open(standard_file_path, 'wb+') as destination:
                            for chunk in standard_file.chunks():
                                destination.write(chunk)
                        
                        # 更新标准图片路径字段，如果模型中有的话
                        if hasattr(hanzi_instance, 'standard_image_path'):
                            hanzi_instance.standard_image_path = f"uploads/{standard_filename}"
                    elif new_structure != original_structure and hasattr(hanzi, 'standard_image_path') and hanzi.standard_image_path:
                        # 处理标准图片路径更新
                        old_standard_path = os.path.join(UPLOAD_FOLDER, os.path.basename(hanzi.standard_image_path))
                        if os.path.exists(old_standard_path):
                            # 获取文件扩展名
                            filename_parts = os.path.basename(hanzi.standard_image_path).split('.')
                            if len(filename_parts) > 1:
                                ext = filename_parts[-1]
                                # 新文件名使用新ID
                                new_standard_filename = f"{hanzi_instance.id}_standard.{ext}"
                                new_standard_path = os.path.join(UPLOAD_FOLDER, new_standard_filename)
                                # 复制文件
                                shutil.copy2(old_standard_path, new_standard_path)
                                # 删除旧文件
                                os.remove(old_standard_path)
                                logger.info(f"已重命名标准图片文件: {os.path.basename(old_standard_path)} -> {new_standard_filename}")
                                # 更新数据库中的标准图片路径
                                hanzi_instance.standard_image_path = f"uploads/{new_standard_filename}"
                    
                    # 如果结构变化导致ID变化，需要删除旧记录并创建新记录
                    if new_structure != original_structure:
                        # 删除旧记录
                        Hanzi.objects.filter(id=hanzi_id).delete()
                        # 重新创建新记录
                        hanzi_instance.save()
                    else:
                        # 正常保存
                        hanzi_instance.save()
                    
                    # 重定向到详情页，保留返回参数
                    return redirect(back_url)
            
            except Exception as e:
                # 处理异常
                logger.error(f"保存汉字时出错: {e}")
                logger.error(traceback.format_exc())
                return render(request, 'hanzi_app/edit.html', {
                    'form': form,
                    'hanzi': hanzi,
                    'error': f"保存失败: {str(e)}",
                    'back_url': back_url,
                    'structure_options': structure_options,
                    'variant_options': variant_options
                })
    else:
        form = HanziForm(instance=hanzi)
    
    return render(request, 'hanzi_app/edit.html', {
        'form': form,
        'hanzi': hanzi,
        'back_url': back_url,
        'structure_options': structure_options,
        'variant_options': variant_options
    })

@csrf_exempt
def update_hanzi(request, hanzi_id):
    if request.method == 'POST':
        try:
            hanzi = get_object_or_404(Hanzi, id=hanzi_id)
            original_structure = hanzi.structure  # 保存原始结构
            
            form = HanziForm(request.POST, request.FILES, instance=hanzi)
            
            # 构建返回URL，添加返回标记并保留所有筛选条件
            url_params = {
                'returning': '1',
            }
            
            # 保留所有筛选参数和分页参数
            filter_params = ['search', 'structure', 'level', 'variant', 'stroke_count', 'page']
            for param in filter_params:
                value = request.GET.get(param)
                if value:
                    url_params[param] = value
            
            # 构建完整的返回URL
            back_url = reverse('hanzi_app:index') + '?' + urlencode(url_params)
            
            if form.is_valid():
                try:
                    # 开启事务处理
                    with transaction.atomic():
                        hanzi_instance = form.save(commit=False)
                        
                        # 检查结构是否发生变化
                        new_structure = hanzi_instance.structure
                        if new_structure != original_structure:
                            # 结构变化，生成新ID
                            old_id = hanzi_instance.id
                            new_id = generate_new_id(new_structure)
                            logger.info(f"汉字'{hanzi_instance.character}'结构从'{original_structure}'变为'{new_structure}'，ID从'{old_id}'更新为'{new_id}'")
                            
                            # 更新ID
                            hanzi_instance.id = new_id
                        
                        # 处理图片上传 - 用户图片
                        if 'new_image_file' in request.FILES:
                            # 先删除旧图片
                            if hanzi.image_path:
                                old_image_path = os.path.join(UPLOAD_FOLDER, os.path.basename(hanzi.image_path))
                                if os.path.exists(old_image_path):
                                    os.remove(old_image_path)
                            
                            image_file = request.FILES['new_image_file']
                            filename = f"{hanzi_instance.id}.{image_file.name.split('.')[-1]}"
                            file_path = os.path.join(UPLOAD_FOLDER, filename)
                            
                            with open(file_path, 'wb+') as destination:
                                for chunk in image_file.chunks():
                                    destination.write(chunk)
                            
                            # 更新图片路径
                            hanzi_instance.image_path = f"uploads/{filename}"
                        elif new_structure != original_structure and hanzi.image_path:
                            # 结构变化但没有上传新图片，需要更新现有图片文件名
                            old_image_path = os.path.join(UPLOAD_FOLDER, os.path.basename(hanzi.image_path))
                            if os.path.exists(old_image_path):
                                # 获取文件扩展名
                                filename_parts = os.path.basename(hanzi.image_path).split('.')
                                if len(filename_parts) > 1:
                                    ext = filename_parts[-1]
                                    # 新文件名使用新ID
                                    new_filename = f"{hanzi_instance.id}.{ext}"
                                    new_image_path = os.path.join(UPLOAD_FOLDER, new_filename)
                                    # 复制文件
                                    shutil.copy2(old_image_path, new_image_path)
                                    # 删除旧文件
                                    os.remove(old_image_path)
                                    logger.info(f"已重命名图片文件: {os.path.basename(old_image_path)} -> {new_filename}")
                                    # 更新数据库中的图片路径
                                    hanzi_instance.image_path = f"uploads/{new_filename}"
                        
                        # 处理标准图片上传
                        if 'new_standard_file' in request.FILES:
                            # 标准图片处理逻辑
                            standard_file = request.FILES['new_standard_file']
                            standard_filename = f"{hanzi_instance.id}_standard.{standard_file.name.split('.')[-1]}"
                            standard_file_path = os.path.join(UPLOAD_FOLDER, standard_filename)
                            
                            with open(standard_file_path, 'wb+') as destination:
                                for chunk in standard_file.chunks():
                                    destination.write(chunk)
                            
                            # 更新标准图片路径字段，如果模型中有的话
                            if hasattr(hanzi_instance, 'standard_image_path'):
                                hanzi_instance.standard_image_path = f"uploads/{standard_filename}"
                        elif new_structure != original_structure and hasattr(hanzi, 'standard_image_path') and hanzi.standard_image_path:
                            # 处理标准图片路径更新
                            old_standard_path = os.path.join(UPLOAD_FOLDER, os.path.basename(hanzi.standard_image_path))
                            if os.path.exists(old_standard_path):
                                # 获取文件扩展名
                                filename_parts = os.path.basename(hanzi.standard_image_path).split('.')
                                if len(filename_parts) > 1:
                                    ext = filename_parts[-1]
                                    # 新文件名使用新ID
                                    new_standard_filename = f"{hanzi_instance.id}_standard.{ext}"
                                    new_standard_path = os.path.join(UPLOAD_FOLDER, new_standard_filename)
                                    # 复制文件
                                    shutil.copy2(old_standard_path, new_standard_path)
                                    # 删除旧文件
                                    os.remove(old_standard_path)
                                    logger.info(f"已重命名标准图片文件: {os.path.basename(old_standard_path)} -> {new_standard_filename}")
                                    # 更新数据库中的标准图片路径
                                    hanzi_instance.standard_image_path = f"uploads/{new_standard_filename}"
                        
                        # 如果结构变化导致ID变化，需要删除旧记录并创建新记录
                        if new_structure != original_structure:
                            # 删除旧记录
                            Hanzi.objects.filter(id=hanzi_id).delete()
                            # 重新创建新记录
                            hanzi_instance.save()
                        else:
                            # 正常保存
                            hanzi_instance.save()
                        
                        # 返回JSON响应
                        return JsonResponse({
                            'success': True,
                            'message': '汉字更新成功',
                            'id': hanzi_instance.id,
                            'back_url': back_url
                        })
                        
                except Exception as e:
                    # 记录错误
                    logger.error(f"保存汉字时出错: {e}")
                    logger.error(traceback.format_exc())
                    return JsonResponse({
                        'success': False,
                        'error': f"保存失败: {str(e)}"
                    }, status=500)
            else:
                # 表单验证失败
                errors = dict([(k, [str(e) for e in v]) for k, v in form.errors.items()])
                return JsonResponse({
                    'success': False,
                    'error': '表单验证失败',
                    'form_errors': errors
                }, status=400)
                
        except Exception as e:
            # 处理异常
            logger.error(f"处理汉字更新请求时出错: {e}")
            logger.error(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'error': f"处理请求失败: {str(e)}"
            }, status=500)
    
    return JsonResponse({'success': False, 'error': '不支持的请求方法'}, status=405)

class EventStreamResponse(StreamingHttpResponse):
    def __init__(self, streaming_content=(), **kwargs):
        super().__init__(streaming_content, content_type='text/event-stream', **kwargs)
        self['Cache-Control'] = 'no-cache'
        self['X-Accel-Buffering'] = 'no'



# process_import_image,get_auto_stroke_order,process_hanzi_data,handle_import_file均为导入数据视图函数import_data的准备处理函数

def process_import_image(image_filename, image_dir, hanzi_id, found_image_path=None):
    """处理导入的图片文件
    
    Args:
        image_filename: 图片文件名
        image_dir: 解压的临时目录
        found_image_path: 如果已知路径，可以直接传入
        
    Returns:
        str: 图片的相对路径，如果未找到则返回空字符串
    """
    if not image_filename:
        return ""
        
    # 确保文件名包含扩展名
    if not image_filename.lower().endswith(('.jpg', '.png')):
        image_filename = f"{image_filename}.jpg"
    
    # 如果已知路径，直接使用
    if found_image_path:
        src_path = found_image_path
        found = True
    else:
        # 查找图片文件
        found = False
        src_path = None
        for root, dirs, files in os.walk(image_dir):
            for file in files:
                if file.lower() == image_filename.lower():
                    src_path = os.path.join(root, file)
                    found = True
                    break
            if found:
                break
    
    if found and src_path:
        # 保存用户图片到uploads目录,用hanzi_id作为文件名，扩展名统一为jpg
        dest_filename = f"{hanzi_id}.jpg"
        dest_path = os.path.join(UPLOAD_FOLDER, dest_filename)
        shutil.copy(src_path, dest_path)
        return f"uploads/{dest_filename}"
    
    return ""

def get_auto_stroke_order(character):
    """获取汉字的自动笔顺
    
    Args:
        character: 汉字字符
        
    Returns:
        str: 笔顺字符串
    """
    if not character:
        return ""
        
    stroke_orders = get_stroke_order(character)
    if stroke_orders and stroke_orders[0]:
        return stroke_orders[0]
    return ""

def process_hanzi_data(data_item, image_dir=None, zip_file=None, is_json=True):
    """处理单个汉字数据项，用于导入
    
    Args:
        data_item: 数据项（字典或DataFrame行）
        image_dir: 解压的临时目录
        zip_file: 上传的ZIP文件
        is_json: 是否是JSON格式(True)或Excel格式(False)
        
    Returns:
        tuple: (成功标志, 汉字对象或错误消息)
    """
    try:
        # 提取汉字字符（必须字段）和ID
        if is_json:
            char = data_item.get('character')
            hanzi_id = str(data_item.get('id', '')).strip()
        else: # Excel处理
            id_key = None
            
            # 尝试根据列名找到ID和字符
            if 'id' in data_item:
                # 确保ID保留前导零
                if isinstance(data_item['id'], (int, float)) and not pd.isna(data_item['id']):
                    # 如果是数字类型，格式化为5位数，保留前导零
                    hanzi_id = f"{int(data_item['id']):05d}"
                else:
                    hanzi_id = str(data_item['id']).strip()
            else:
                hanzi_id = ''
                
            if 'character' in data_item and not pd.isna(data_item['character']):
                char = str(data_item['character']).strip()
                # 检查是否是有效的单个汉字
                if len(char) != 1:
                    # 可能是Excel导入的特殊格式，需要进一步处理
                    pass
            else:
                char = ''
                
            # 如果没有找到有效的字符，尝试从第一行数据中提取
            if not char or len(char) != 1:
                # 查找字典格式的数据
                for key, value in data_item.items():
                    # 跳过空值
                    if pd.isna(key) or pd.isna(value):
                        continue
                        
                    str_key = str(key).strip()
                    str_value = str(value).strip()
                    
                    # 尝试找出ID和汉字
                    if len(str_key) == 5 and str_key.startswith('0'):  # 可能是ID
                        id_key = str_key
                    elif len(str_key) == 1 and '\u4e00' <= str_key <= '\u9fff':  # 是汉字
                        char = str_key
                    elif len(str_value) == 1 and '\u4e00' <= str_value <= '\u9fff':  # 值是汉字
                        char = str_value
                        
                # 如果找到了ID但没有hanzi_id
                if id_key and not hanzi_id:
                    hanzi_id = id_key
                    
            # 额外检查：如果是数字ID，确保格式化为5位带前导零的字符串
            if hanzi_id and hanzi_id.isdigit() and len(hanzi_id) < 5:
                hanzi_id = f"{int(hanzi_id):05d}"
        
        # 如果仍然没有有效的汉字字符
        if not char or len(str(char).strip()) != 1 or not '\u4e00' <= char <= '\u9fff':
            return False, f"记录缺少有效的汉字字符: {data_item}"
        
        char = str(char).strip()
        
        # 确保ID格式正确（5位，带前导零）
        if hanzi_id and hanzi_id.isdigit() and len(hanzi_id) < 5:
            hanzi_id = f"{int(hanzi_id):05d}"
        
        # 获取结构信息
        if is_json:
            structure = data_item.get('structure', '未知结构')
        else:
            structure = data_item.get('structure', '未知结构')
            if pd.isna(structure):
                structure = '未知结构'
        
        # 检查ID是否存在
        existing_hanzi = None
        is_update = False
        
        if hanzi_id:
            try:
                # 尝试匹配ID
                existing_hanzi = Hanzi.objects.get(id=hanzi_id)
                is_update = True
                logger.info(f"找到ID为 {hanzi_id} 的现有记录，将进行更新")
                
                # 检查结构是否变化
                if existing_hanzi.structure != structure:
                    logger.info(f"汉字'{char}'的结构由'{existing_hanzi.structure}'变为'{structure}'，需要更新ID")
                    # 结构变化，生成新ID
                    old_id = hanzi_id
                    hanzi_id = generate_new_id(structure)
                    logger.info(f"为汉字'{char}'生成新ID: {hanzi_id}")
                    
                    # 如果有图片，处理图片文件
                    if existing_hanzi.image_path:
                        old_image_path = os.path.join(UPLOAD_FOLDER, os.path.basename(existing_hanzi.image_path))
                        if os.path.exists(old_image_path):
                            # 获取文件扩展名
                            filename_parts = os.path.basename(existing_hanzi.image_path).split('.')
                            if len(filename_parts) > 1:
                                ext = filename_parts[-1]
                                # 新文件名使用新ID
                                new_filename = f"{hanzi_id}.{ext}"
                                new_image_path = os.path.join(UPLOAD_FOLDER, new_filename)
                                # 复制文件
                                shutil.copy2(old_image_path, new_image_path)
                                # 删除旧文件
                                os.remove(old_image_path)
                                logger.info(f"已重命名图片文件: {os.path.basename(old_image_path)} -> {new_filename}")
                                # 更新数据库中的图片路径
                                image_path = f"uploads/{new_filename}"
                    
                    # 如果有标准图片，处理标准图片文件
                    if hasattr(existing_hanzi, 'standard_image_path') and existing_hanzi.standard_image_path:
                        old_standard_path = os.path.join(UPLOAD_FOLDER, os.path.basename(existing_hanzi.standard_image_path))
                        if os.path.exists(old_standard_path):
                            # 获取文件扩展名
                            filename_parts = os.path.basename(existing_hanzi.standard_image_path).split('.')
                            if len(filename_parts) > 1:
                                ext = filename_parts[-1]
                                # 新文件名使用新ID
                                new_standard_filename = f"{hanzi_id}_standard.{ext}"
                                new_standard_path = os.path.join(UPLOAD_FOLDER, new_standard_filename)
                                # 复制文件
                                shutil.copy2(old_standard_path, new_standard_path)
                                # 删除旧文件
                                os.remove(old_standard_path)
                                logger.info(f"已重命名标准图片文件: {os.path.basename(old_standard_path)} -> {new_standard_filename}")
                    
                    # 将更新模式设为False，因为我们将删除旧记录并创建新记录
                    is_update = False
                    
                    # 删除旧记录
                    existing_hanzi.delete()
                    logger.info(f"已删除原ID为{old_id}的记录")
                    existing_hanzi = None
                
            except Hanzi.DoesNotExist:
                if not existing_hanzi:
                    # 如果没有找到匹配的ID，则创建新记录
                    logger.info(f"未找到ID为 {hanzi_id} 的记录，将创建新记录")
        else:
            # 没有提供ID，生成新ID
            hanzi_id = generate_new_id(structure)
            logger.info(f"未提供ID，生成新ID: {hanzi_id}")
          
        
        # 处理图片文件
        image_path = ""
        if zip_file:
            if is_json and 'image_path' in data_item and data_item['image_path']:
                image_path = process_import_image(data_item['image_path'], image_dir, hanzi_id)
            elif not is_json:
                # 处理Excel中的image_path
                if 'image_path' in data_item and not pd.isna(data_item['image_path']):
                    image_filename = str(data_item['image_path'])
                    image_path = process_import_image(image_filename, image_dir, hanzi_id)
                # 如果未找到，尝试查找格式为A开头的值作为图片路径
                else:
                    for key, value in data_item.items():
                        if pd.isna(value):
                            continue
                        str_value = str(value).strip()
                        if str_value.startswith('A') and str_value[1:].isdigit():
                            image_path = process_import_image(str_value, image_dir, hanzi_id)
                            break
        
        # 如果检验发现是更新汉字，尝试使用现有图片路径（is_update=True）
        if is_update and not image_path and existing_hanzi.image_path:
            image_path = existing_hanzi.image_path
            
        # 生成笔顺
        stroke_order = get_auto_stroke_order(char)
            
        # 生成拼音
        pinyin = get_pinyin(char)[0] if get_pinyin(char) else ''
            
        # 生成笔画数
        stroke_count = stroke_dict.get(char, 0)
        
        # 获取其他字段
        if is_json:
            level = data_item.get('level', 'D')
            variant = data_item.get('variant', '简体')
            comment = data_item.get('comment', '无')
        else:
            level = data_item.get('level', 'D')
            if pd.isna(level):
                level = 'D'
                
            variant = data_item.get('variant', '简体')
            if pd.isna(variant):
                variant = '简体'
                
            comment = data_item.get('comment', '无')
            if pd.isna(comment):
                comment = '无'
                
        # 准备汉字数据
        hanzi_data = {
            'id': hanzi_id,
            'character': char,
            'stroke_count': stroke_count,
            'structure': structure,
            'pinyin': pinyin,
            'level': level,
            'variant': variant,
            'comment': comment,
            'image_path': image_path,
            'stroke_order': stroke_order
        }
        
        # 创建或更新汉字对象
        if is_update:
            # 更新记录
            for key, value in hanzi_data.items():
                setattr(existing_hanzi, key, value)
            hanzi_obj = existing_hanzi
            logger.info(f"更新汉字记录: ID={hanzi_id}, 字符={char}")
        else:
            # 创建记录
            hanzi_obj = Hanzi(**hanzi_data)
            logger.info(f"创建新汉字记录: ID={hanzi_id}, 字符={char}")
            
        # 保存到数据库
        hanzi_obj.save()
        
        # 自动生成标准图片并
        standard_path = generate_hanzi_image(char)
        if standard_path:
            # 提取相对路径并更新数据库
            rel_path = os.path.join("standard_images", f"{char}.jpg")
            Hanzi.objects.filter(id=hanzi_id).update(standard_image=rel_path)
            
        return True, hanzi_obj
        
    except Exception as e:
        error_msg = f"处理记录时出错: {str(e)}, 记录: {data_item}"
        logger.error(error_msg)
        return False, error_msg

def handle_import_file(file_obj, zip_file=None, is_json=True):
    """处理导入文件并返回事件生成器
    
    Args:
        file_obj: 文件对象（JSON或Excel）
        zip_file: ZIP压缩文件（可选）
        is_json: 是否为JSON文件（True）或Excel文件（False）
        
    Returns:
        generator: 事件生成器函数
    """
    # 创建临时解压目录
    image_dir = os.path.join(settings.MEDIA_ROOT, 'temp_images', str(time.time()))
    os.makedirs(image_dir, exist_ok=True)
    
    # 如果有ZIP文件，解压它
    if zip_file:
        with zipfile.ZipFile(zip_file, 'r') as zf:
            zf.extractall(image_dir)
            
    def generate_events():
        try:
            # 解析文件
            if is_json:
                data = json.loads(file_obj.read().decode('utf-8'))
            else:
                # 尝试不同的header选项，找到合适的解析方式
                try:
                    # 首先尝试使用第一行作为列名
                    df = pd.read_excel(file_obj)
                    # 检查解析后的数据结构是否合理
                    if 'id' not in df.columns and 'character' not in df.columns:
                        # 如果没有正确的列名，尝试使用第二行作为列名
                        df = pd.read_excel(file_obj, header=1)
                except Exception as e:
                    logger.warning(f"Excel解析失败，尝试备用方法: {str(e)}")
                    df = pd.read_excel(file_obj, header=None)
                    # 尝试检测并处理列名
                    if df.shape[0] > 0:
                        first_row = df.iloc[0]
                        if any(str(col).lower() == 'id' for col in first_row) or any(str(col).lower() == 'character' for col in first_row):
                            df.columns = first_row
                            df = df.iloc[1:]
                
                # 转换为字典列表
                data = df.to_dict('records')
                
            total = len(data)
            processed = 0
            success_count = 0
            errors = []
            
            for item in data:
                success, result = process_hanzi_data(item, image_dir, zip_file, is_json)
                
                if success:
                    success_count += 1
                else:
                    errors.append(result)
                
                processed += 1
                progress = (processed / total * 100)
                yield f"data: {json.dumps({'progress': progress, 'processed': processed, 'success': success_count, 'errors': len(errors)})}\n\n"
            
            # 导入完成后返回结果
            yield f"data: {json.dumps({'success': True, 'message': f'成功导入 {success_count} 条记录，失败 {len(errors)} 条', 'errors': errors})}\n\n"
            
            # 导入完成后自动生成所有标准图片
            from .generate import generate_all_standard_images
            generate_all_standard_images()
            
            # 清理临时目录
            try:
                shutil.rmtree(image_dir)
            except Exception as e:
                logger.error(f"清理临时目录失败: {str(e)}")
                
        except Exception as e:
            logger.error(f"导入失败: {str(e)}")
            yield f"data: {json.dumps({'success': False, 'message': f'导入失败: {str(e)}'})}\n\n"
            
    return generate_events

# 导入数据视图函数
@require_http_methods(["GET", "POST"])
def import_data(request):
    if request.method == 'POST':
        try:
            # 处理Excel文件导入
            if 'excel_file' in request.FILES:
                excel_file = request.FILES['excel_file']
                zip_file = request.FILES.get('image_zip')
                
                # 使用通用处理函数
                events_generator = handle_import_file(excel_file, zip_file, is_json=False)
                return EventStreamResponse(events_generator())
            
            # 处理JSON文件导入
            elif 'file' in request.FILES:
                json_file = request.FILES['file']
                zip_file = request.FILES.get('image_zip')
                
                # 使用通用处理函数
                events_generator = handle_import_file(json_file, zip_file, is_json=True)
                return EventStreamResponse(events_generator())
            
            else:
                return JsonResponse({'success': False, 'message': '未选择文件'})
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return render(request, 'hanzi_app/import.html')

def export_hanzi(request):
    """导出汉字数据"""
    if request.method == 'POST':
        # 从POST获取用户选择的字段和图片选项
        selected_fields = request.POST.getlist('fields', [])
        
        # 确保character字段始终被包含
        if 'character' not in selected_fields:
            selected_fields.append('character')
        
        # 至少需要选择一个字段
        if not selected_fields:
            messages.error(request, "请至少选择一个导出字段")
            return redirect(request.path)
        
        # 获取图片选项
        include_images = 'include_images' in request.POST
        include_standard_images = 'include_standard_images' in request.POST
        embed_images_in_excel = 'embed_images_in_excel' in request.POST
        
        # 如果要在Excel中嵌入图片，需要确保图片字段被选中
        if embed_images_in_excel:
            if include_images and 'image_path' not in selected_fields:
                selected_fields.append('image_path')
            if include_standard_images and 'standard_image' not in selected_fields:
                selected_fields.append('standard_image')
        
        # 从会话获取筛选参数
        export_filters = request.session.get('export_filters', {})
        
        # 获取筛选后的汉字列表
        filtered_hanzi = get_filtered_hanzi_list(
            export_filters.get('search', ''),
            export_filters.get('structure', ''),
            export_filters.get('level', ''),
            export_filters.get('variant', ''),
            export_filters.get('stroke_count', ''),
            export_filters.get('ids', ''),
            export_filters.get('stroke_count_min', ''),
            export_filters.get('stroke_count_max', '')
        )
        
        # 生成导出文件
        export_files = generate_export_files(
            filtered_hanzi, 
            include_images,
            include_standard_images,
            selected_fields,
            embed_images_in_excel
        )
        export_files['count'] = filtered_hanzi.count()  # 添加记录数量
        
        # 保存导出文件信息到会话
        request.session['export_files'] = export_files
        request.session['export_timestamp'] = int(time.time())
        
        # 构建导出页面URL
        export_page_url = reverse('hanzi_app:export_page')
        
        # 获取返回URL
        return_url = request.POST.get('return_url', reverse('hanzi_app:index'))
        
        # 构建完整URL
        if return_url:
            params = {'return_url': return_url}
            export_page_url += '?' + urlencode(params)
        
        # 重定向到导出页面
        return redirect(export_page_url)
    else:
        # 直接在export_hanzi函数中处理GET请求，显示导出选项页面
        # 获取筛选条件
        ids = request.GET.get('ids', '')
        search = request.GET.get('search', '')
        structure = request.GET.get('structure', '')
        level = request.GET.get('level', '')
        variant = request.GET.get('variant', '')
        stroke_count = request.GET.get('stroke_count', '')
        stroke_count_min = request.GET.get('stroke_count_min', '')
        stroke_count_max = request.GET.get('stroke_count_max', '')
        return_url = request.GET.get('return_url', reverse('hanzi_app:index'))
        
        # 获取筛选后的汉字列表
        filtered_hanzi = get_filtered_hanzi_list(
            search, structure, level, variant, stroke_count, 
            ids, stroke_count_min, stroke_count_max
        )
        
        # 保存筛选参数到会话，供后续使用
        request.session['export_filters'] = {
            'ids': ids,
            'search': search,
            'structure': structure,
            'level': level,
            'variant': variant,
            'stroke_count': stroke_count,
            'stroke_count_min': stroke_count_min,
            'stroke_count_max': stroke_count_max
        }
        
        # 构建筛选信息对象，用于显示在前端
        filter_info = {
            'structure': structure if structure and structure != '所有' else None,
            'variant': variant if variant and variant != '所有' else None,
            'level': level if level and level != '所有' else None,
            'search_term': search if search else None,
            'stroke_count': stroke_count if stroke_count and stroke_count != '所有' else None,
        }
        
        # 可选字段列表
        available_fields = [
            {'id': 'id', 'name': '编号'},
            {'id': 'character', 'name': '汉字', 'checked': True},
            {'id': 'structure', 'name': '结构', 'checked': True},
            {'id': 'variant', 'name': '简繁体', 'checked': True},
            {'id': 'level', 'name': '等级', 'checked': True},
            {'id': 'stroke_count', 'name': '笔画数'},
            {'id': 'pinyin', 'name': '拼音'},
            {'id': 'stroke_order', 'name': '笔顺'},
            {'id': 'comment', 'name': '评价', 'checked': True},
            {'id': 'image_path', 'name': '图片路径', 'checked': True}
        ]
        
        # 图片选项
        image_options = [
            {'id': 'include_images', 'name': '包含手写图片'},
            {'id': 'include_standard_images', 'name': '包含标准楷体图片'}
        ]
        
        context = {
            'available_fields': available_fields,
            'image_options': image_options,
            'filtered_hanzi_count': filtered_hanzi.count(),
            'filter_info': filter_info,
            'return_url': return_url
        }
        
        return render(request, 'hanzi_app/export_options.html', context)

# 导出数据视图函数
def export_page(request):
    """导出页面"""
    # 获取会话中的导出文件信息
    export_files = request.session.get('export_files', {})
    export_timestamp = request.session.get('export_timestamp', 0)
    return_url = request.GET.get('return_url', reverse('hanzi_app:index'))
    
    # 检查文件是否存在
    export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
    
    # 准备文件列表
    file_list = []
    
    for file_type in ['json', 'excel', 'zip']:
        if file_type in export_files:
            file_path = os.path.join(export_dir, export_files[file_type])
            if os.path.exists(file_path):
                # 文件存在，添加到文件列表
                file_info = {
                    'name': export_files[file_type],
                    'size': f"{os.path.getsize(file_path) / (1024 * 1024):.2f} MB",
                    'type': '数据文件' if file_type == 'json' else ('表格文件' if file_type == 'excel' else '图片压缩包'),
                    'icon': 'fa-file-code' if file_type == 'json' else ('fa-file-excel' if file_type == 'excel' else 'fa-file-archive')
                }
                file_list.append(file_info)
    
    # 检查是否有导出数据
    export_count = export_files.get('count', 0)
    
    # 如果没有记录，显示警告信息
    if export_count == 0:
        messages.warning(request, "没有符合筛选条件的汉字数据")
    
    context = {
        'export_files': export_files,
        'export_timestamp': datetime.fromtimestamp(export_timestamp) if export_timestamp else None,
        'return_url': return_url,
        'current_time': int(time.time()),
        'file_list': file_list,
        'export_count': export_count
    }
    
    return render(request, 'hanzi_app/export.html', context)

def delete_export_file(request, filename):
    """删除单个导出文件"""
    try:
        if not filename:
            return JsonResponse({'error': '未指定文件名'}, status=400)
        
        file_path = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"已删除导出文件: {filename}")
            
            # 更新session中的文件信息
            export_files = request.session.get('export_files', {})
            for key in ['json_file', 'excel_file', 'zip_file']:
                if key in export_files and export_files[key] == filename:
                    del export_files[key]
                    request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': f'已删除文件: {filename}'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': '文件不存在'
            })
    except Exception as e:
        logger.error(f"删除文件失败: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def get_filtered_hanzi_list(search, structure, level, variant, stroke_count, ids, stroke_count_min, stroke_count_max):
    """根据筛选条件获取汉字列表"""
    queryset = Hanzi.objects.all()
    
    # 如果指定了ID列表，则使用ID列表筛选
    if ids:
        id_list = ids.split(',')
        queryset = queryset.filter(id__in=id_list)
        return queryset
    
    # 否则使用其他筛选条件
    if search:
        queryset = queryset.filter(
            Q(id__icontains=search) | 
            Q(character__icontains=search) | 
            Q(pinyin__icontains=search) | 
            Q(comment__icontains=search)
        )
    
    if structure and structure != '所有':
        queryset = queryset.filter(structure=structure)
    
    if level and level != '所有':
        queryset = queryset.filter(level=level)
    
    if variant and variant != '所有':
        queryset = queryset.filter(variant=variant)
    
    if stroke_count and stroke_count != '所有':
        try:
            stroke_count_int = int(stroke_count)
            queryset = queryset.filter(stroke_count=stroke_count_int)
        except ValueError:
            pass
    
    if stroke_count_min and stroke_count_max:
        try:
            stroke_count_min_int = int(stroke_count_min)
            stroke_count_max_int = int(stroke_count_max)
            queryset = queryset.filter(stroke_count__range=(stroke_count_min_int, stroke_count_max_int))
        except ValueError:
            pass
    
    return queryset

def generate_export_files(filtered_hanzi, include_images=False, include_standard_images=False, selected_fields=[], embed_images_in_excel=False):
    """生成导出文件"""
    try:
        # 创建导出目录
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
        
        # 时间戳
        timestamp = int(time.time())
        file_prefix = f'hanzi_export_{timestamp}'
        
        # 确保至少包含字符字段
        if 'character' not in selected_fields:
            selected_fields.append('character')
        
        # 构建返回结果
        result = {
            'timestamp': timestamp,
            'count': filtered_hanzi.count(),
            'filter_info': {
                'total_hanzi': Hanzi.objects.count(),
                'filtered_hanzi': filtered_hanzi.count()
            },
            'selected_fields': selected_fields
        }
        
        # 导出JSON文件
        json_file_path = os.path.join(export_dir, f'{file_prefix}.json')
        json_data = custom_serialize_hanzi(filtered_hanzi, selected_fields)
        with open(json_file_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        result['json'] = os.path.basename(json_file_path)
        
        # 导出Excel文件
        excel_file_path = os.path.join(export_dir, f'{file_prefix}.xlsx')
        export_to_excel(filtered_hanzi, excel_file_path, selected_fields, embed_images_in_excel)
        result['excel'] = os.path.basename(excel_file_path)
        
        # 如果包含图片，创建ZIP文件
        if include_images or include_standard_images:
            zip_file_path = os.path.join(export_dir, f'{file_prefix}_with_images.zip')
            with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                # 添加JSON文件
                zipf.write(json_file_path, os.path.basename(json_file_path))
                
                # 添加Excel文件
                zipf.write(excel_file_path, os.path.basename(excel_file_path))
                
                # 添加图片文件
                images_added = 0
                
                # 添加手写图片
                if include_images:
                    for hanzi in filtered_hanzi:
                        if hanzi.image_path:
                            image_path = os.path.join(settings.MEDIA_ROOT, hanzi.image_path)
                            if os.path.exists(image_path):
                                # 获取不带后缀的文件名
                                filename = os.path.basename(hanzi.image_path)
                                base_filename = os.path.splitext(filename)[0] 
                                # 使用不带后缀的文件名作为目标文件名
                                zipf.write(image_path, f'images/{base_filename}' + os.path.splitext(image_path)[1])
                                images_added += 1
                
                # 添加标准楷体图片
                if include_standard_images:
                    for hanzi in filtered_hanzi:
                        if hanzi.standard_image:
                            standard_image_path = os.path.join(settings.MEDIA_ROOT, hanzi.standard_image)
                            if os.path.exists(standard_image_path):
                                # 获取不带后缀的文件名
                                filename = os.path.basename(hanzi.standard_image)
                                base_filename = os.path.splitext(filename)[0]
                                # 使用不带后缀的文件名作为目标文件名
                                zipf.write(standard_image_path, f'images/standard_{base_filename}' + os.path.splitext(standard_image_path)[1])
                                images_added += 1
                
                result['zip'] = os.path.basename(zip_file_path)
                result['images_included'] = images_added
        
        return result
    except Exception as e:
        logger.error(f"导出文件生成失败: {str(e)}")
        return {'error': str(e)}

def custom_serialize_hanzi(queryset, fields):
    """自定义序列化汉字对象，根据选择的字段"""
    result = []
    for hanzi in queryset:
        item = {}
        for field in fields:
            if field == 'image_path' and hanzi.image_path:
                # 处理图片路径，只返回文件名部分（不含后缀）
                filename = os.path.basename(hanzi.image_path)
                item[field] = os.path.splitext(filename)[0]
            else:
                # 获取其他字段值
                value = getattr(hanzi, field, '')
                # 对于可能为None的值，提供默认空字符串
                item[field] = value if value is not None else ''
        result.append(item)
    return result

def export_to_excel(filtered_hanzi, excel_file_path, selected_fields, embed_images_in_excel=False):
    """将汉字数据导出到Excel文件，并可选择将图片嵌入到Excel中"""
    try:
        # 创建一个DataFrame来存储数据
        data = []
        for hanzi in filtered_hanzi:
            item = {}
            for field in selected_fields:
                if field == 'image_path' and hanzi.image_path:
                    # 处理图片路径，只保留文件名部分（不含后缀）
                    filename = os.path.basename(hanzi.image_path)
                    item[field] = os.path.splitext(filename)[0]  # 不含后缀的文件名
                elif field == 'standard_image' and hanzi.standard_image:
                    # 处理标准图片路径
                    filename = os.path.basename(hanzi.standard_image)
                    item[field] = os.path.splitext(filename)[0]  # 不含后缀的文件名
                else:
                    # 获取其他字段值
                    value = getattr(hanzi, field, '')
                    # 对于可能为None的值，提供默认空字符串
                    item[field] = value if value is not None else ''
            data.append(item)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 如果不需要嵌入图片，使用简单的Pandas导出
        if not embed_images_in_excel:
            with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='汉字数据')
                
                # 调整列宽
                worksheet = writer.sheets['汉字数据']
                for idx, col in enumerate(df.columns):
                    column_width = max(df[col].astype(str).map(len).max(), len(col) + 2)
                    worksheet.column_dimensions[chr(65 + idx)].width = column_width
            
            return True
        
        # 需要嵌入图片时，使用openpyxl直接操作
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = '汉字数据'
        
        # 检查是否需要添加图片列
        has_image_column = 'image_path' in selected_fields
        has_standard_image_column = 'standard_image' in selected_fields
        
        # 创建新的列表，包含原始字段和额外的图片列
        all_columns = list(selected_fields)
        
        # 在原字段列表后添加图片列
        if has_image_column:
            all_columns.append('手写图片')
        if has_standard_image_column:
            all_columns.append('标准楷体图片')
        
        # 添加表头
        for col_idx, column_name in enumerate(all_columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # 添加数据行
        for row_idx, hanzi in enumerate(filtered_hanzi, 2):
            # 添加常规数据列
            col_idx = 1
            for field in selected_fields:
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # 获取字段值
                value = getattr(hanzi, field, '')
                
                # 对于图片路径字段，只显示文件名部分（不含路径和扩展名）
                if field in ['image_path', 'standard_image'] and value:
                    filename = os.path.basename(value)
                    cell.value = os.path.splitext(filename)[0]  # 不含后缀的文件名
                elif value is not None:
                    cell.value = value
                    
                col_idx += 1
            
            # 添加手写图片列
            if has_image_column:
                if hanzi.image_path:
                    try:
                        img_path = os.path.join(settings.MEDIA_ROOT, hanzi.image_path)
                        if os.path.exists(img_path):
                            img = openpyxl.drawing.image.Image(img_path)
                            # 设置图片大小
                            img.width = 100
                            img.height = 100
                            # 计算图片位置 - 使用新的图片列
                            cell_address = f"{chr(64 + col_idx)}{row_idx}"
                            img.anchor = cell_address
                            worksheet.add_image(img)
                    except Exception as e:
                        logger.error(f"添加手写图片失败: {str(e)}")
                col_idx += 1
            
            # 添加标准图片列
            if has_standard_image_column:
                if hanzi.standard_image:
                    try:
                        img_path = os.path.join(settings.MEDIA_ROOT, hanzi.standard_image)
                        if os.path.exists(img_path):
                            img = openpyxl.drawing.image.Image(img_path)
                            img.width = 100
                            img.height = 100
                            # 计算图片位置 - 使用新的图片列
                            cell_address = f"{chr(64 + col_idx)}{row_idx}"
                            img.anchor = cell_address
                            worksheet.add_image(img)
                    except Exception as e:
                        logger.error(f"添加标准图片失败: {str(e)}")
                col_idx += 1
        
        # 调整列宽和行高
        original_cols_count = len(selected_fields)
        for col_idx, column_name in enumerate(all_columns, 1):
            column_letter = chr(64 + col_idx)
            
            # 如果是图片列（额外添加的列），设置较大的宽度
            if col_idx > original_cols_count:
                worksheet.column_dimensions[column_letter].width = 20
            else:
                # 其他列根据内容自动调整宽度
                max_length = 0
                for cell in worksheet[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 所有图片列共用一种行高设置
        if has_image_column or has_standard_image_column:
            for row_idx in range(2, len(filtered_hanzi) + 2):
                worksheet.row_dimensions[row_idx].height = 75
        
        # 保存Excel文件
        workbook.save(excel_file_path)
        
        return True
    except Exception as e:
        logger.error(f"Excel导出失败: {str(e)}")
        return False

def download_file(request, filename):
    """处理文件下载，并设置下载后删除文件标记"""
    if not filename:
        return JsonResponse({'error': '未指定文件名'}, status=400)
    
    file_path = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
    if not os.path.exists(file_path):
        return JsonResponse({'error': '文件不存在'}, status=404)
    
    try:
        # 使用FileResponse自动处理文件关闭
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
        
        # 设置正确的Content-Type
        if filename.endswith('.json'):
            response['Content-Type'] = 'application/json'
        elif filename.endswith('.zip'):
            response['Content-Type'] = 'application/zip'
        elif filename.endswith('.xlsx'):
            response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            response['Content-Type'] = 'application/octet-stream'
        
        # 记录下载标记，但不立即删除文件
        request.session[f'downloaded_{filename}'] = True
        
        return response
    except Exception as e:
        logger.error(f"文件下载失败: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def cleanup_exports(request):
    """清理所有导出文件"""
    try:
        # 删除所有导出文件夹中的文件
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        if os.path.exists(export_dir):
            deleted_count = 0
            for file in os.listdir(export_dir):
                file_path = os.path.join(export_dir, file)
                if os.path.isfile(file_path):
                    try:
                        os.remove(file_path)
                        deleted_count += 1
                        logger.info(f"已删除临时导出文件: {file}")
                    except Exception as file_error:
                        logger.error(f"删除文件 {file} 失败: {str(file_error)}")
        
        return JsonResponse({
            'success': True, 
            'message': f'已清理导出文件夹，删除了 {deleted_count} 个文件'
        })
    except Exception as e:
        logger.error(f"清理导出文件失败: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)

def clear_selected(request):
    """清除所有选中的条目"""
    return JsonResponse({'success': True})

def generate_new_id(structure):
    """根据结构生成新的汉字ID"""
    structure_map = {
        "未知结构": "0",
        "左右结构": "1",
        "上下结构": "2",
        "包围结构": "3",
        "独体结构": "4",
        "品字结构": "5",
        "穿插结构": "6"
    }
    
    prefix = structure_map.get(structure, "0")
    
    # 找到该结构的最大ID
    max_id_obj = Hanzi.objects.filter(id__startswith=prefix).order_by('-id').first()
    
    if max_id_obj:
        # 从现有ID中提取数字部分并递增
        try:
            current_num = int(max_id_obj.id[1:])
            new_num = current_num + 1
            new_id = f"{prefix}{new_num:04d}"
        except ValueError:
            # 如果现有ID不符合预期格式，则生成新的
            new_id = f"{prefix}0001"
    else:
        # 如果没有该结构的汉字，从1开始
        new_id = f"{prefix}0001"
    
    return new_id

@cache_page(60 * 15)  # 缓存15分钟
def get_stroke_order_api(request, char):
    # 使用缓存键
    cache_key = f'stroke_order_{char}'
    stroke_order = cache.get(cache_key)
    
    if stroke_order is None:
        # 缓存未命中，从数据库或文件获取
        from .generate import get_stroke_order
        stroke_orders = get_stroke_order(char)
        if stroke_orders and stroke_orders[0]:
            # 去除中括号和引号
            stroke_order = stroke_orders[0].strip("[]'")
        else:
            stroke_order = ''
        # 存入缓存，有效期1天
        cache.set(cache_key, stroke_order, 60*60*24)
    
    return JsonResponse({'stroke_order': stroke_order})

# 添加笔顺搜索视图
def stroke_search(request):
    """笔顺搜索页面"""
    stroke_pattern = request.GET.get('stroke_pattern', '')
    results = []
    
    # 常用笔画列表，用于快速选择
    common_strokes = ['横','竖','点','撇','横折',
                      '捺','横撇/横钩','提','横折钩','撇折',
                      '竖钩','竖弯钩','竖折/竖弯','竖提','斜钩',
                      '撇点','竖折折钩','横折弯钩/横斜钩','横折折折钩/横撇弯钩','横折折撇',
                      '弯钩','横折折/横折弯','横折提','竖折撇/竖折折','横折折折']

    if stroke_pattern:
        # 执行搜索
        results = Hanzi.search_by_stroke_order(stroke_pattern)
        
        # 分页处理
        paginator = Paginator(results, 24)  # 每页显示25条
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # 为每个汉字添加动画延迟值
        for i, hanzi in enumerate(page_obj.object_list):
            hanzi.animation_delay = (i + 8) * 50
    else:
        page_obj = None
    
    context = {
        'stroke_pattern': stroke_pattern,
        'common_strokes': common_strokes,
        'page_obj': page_obj,
        'results_count': len(results) if results else 0
    }
    
    return render(request, 'hanzi_app/stroke_search.html', context)

def import_view(request):
    """导入数据页面入口"""
    return render(request, 'hanzi_app/import.html', {
        'title': '数据导入'
    })

def update_task_status(status_file, progress, message, status=None, result_file=None):
    """更新任务状态文件"""
    try:
        with open(status_file, 'r', encoding='utf-8') as f:
            task_status = json.load(f)
        
        # 更新字段
        task_status['progress'] = progress
        task_status['message'] = message
        
        if status:
            task_status['status'] = status
        
        if result_file:
            task_status['result_file'] = result_file
        
        with open(status_file, 'w', encoding='utf-8') as f:
            json.dump(task_status, f)
            
        logger.info(f"任务状态更新: {progress}% - {message}")
    except Exception as e:
        logger.error(f"更新任务状态失败: {str(e)}")


def get_completed_files(output_dir):
    """获取已完成的导入文件列表"""
    completed_files = []
    
    try:
        # 确保目录存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"创建导入结果目录: {output_dir}")
            return completed_files
        
        # 获取所有Excel文件和日志文件
        all_files = os.listdir(output_dir)
        excel_files = [f for f in all_files if f.lower().endswith('.xlsx') and 'hanzi_import_' in f.lower()]
        log_files = [f for f in all_files if f.lower().endswith('_failed.log') and 'hanzi_import_' in f.lower()]
        
        # 创建日志文件映射
        log_map = {}
        for log_file in log_files:
            # 从日志文件名提取基本名称（不包括_failed.log后缀）
            base_name = os.path.splitext(log_file)[0].replace('_failed', '')
            log_map[base_name] = log_file
        
        # 按修改时间排序（最新的在前）
        excel_files.sort(key=lambda f: os.path.getmtime(os.path.join(output_dir, f)), reverse=True)
        
        # 提取文件信息
        for filename in excel_files:
            file_path = os.path.join(output_dir, filename)
            if os.path.isfile(file_path):
                # 构建文件URL
                if output_dir.startswith(settings.MEDIA_ROOT):
                    relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
                    file_url = f"/media/{relative_path.replace(os.sep, '/')}"
                else:
                    file_url = f"/media/import_results/{filename}"
                
                # 获取文件修改时间和大小
                mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                formatted_time = mod_time.strftime('%Y-%m-%d %H:%M:%S')
                file_size = os.path.getsize(file_path)
                
                # 检查是否有对应的日志文件
                base_name = os.path.splitext(filename)[0]
                log_filename = f"{base_name}_failed.log"
                log_path = os.path.join(output_dir, log_filename)
                has_log = os.path.exists(log_path)
                
                # 构建日志文件URL
                log_url = None
                if has_log:
                    if output_dir.startswith(settings.MEDIA_ROOT):
                        relative_log_path = os.path.relpath(log_path, settings.MEDIA_ROOT)
                        log_url = f"/media/{relative_log_path.replace(os.sep, '/')}"
                    else:
                        log_url = f"/media/import_results/{log_filename}"
                
                # 添加到文件列表
                completed_files.append({
                    'filename': filename,
                    'url': file_url,
                    'timestamp': formatted_time,
                    'size': file_size,
                    'size_formatted': format_file_size(file_size),
                    'has_log': has_log,
                    'log_url': log_url,
                    'log_filename': log_filename if has_log else None
                })
    except Exception as e:
        logger.error(f"获取已完成文件列表失败: {str(e)}")
    
    return completed_files

def format_file_size(size_in_bytes):
    """格式化文件大小"""
    if size_in_bytes < 1024:
        return f"{size_in_bytes} B"
    elif size_in_bytes < 1024 * 1024:
        return f"{size_in_bytes / 1024:.1f} KB"
    elif size_in_bytes < 1024 * 1024 * 1024:
        return f"{size_in_bytes / (1024 * 1024):.1f} MB"
    else:
        return f"{size_in_bytes / (1024 * 1024 * 1024):.1f} GB"

def get_active_tasks(output_dir):
    """获取当前活动的导入任务列表"""
    active_tasks = []
    
    try:
        # 获取所有Celery任务
        inspector = Inspect(app=celery_app)
        reserved = inspector.reserved() or {}
        active = inspector.active() or {}
        
        # 合并任务列表，只保留导入任务
        all_tasks = []
        for worker, tasks in reserved.items():
            for task in tasks:
                if task.get('name') == 'hanzi_app.tasks.process_import_data_task':
                    all_tasks.append((worker, task))
                    
        for worker, tasks in active.items():
            for task in tasks:
                if task.get('name') == 'hanzi_app.tasks.process_import_data_task':
                    all_tasks.append((worker, task))
        
        # 提取任务信息
        for worker_name, task in all_tasks:
            task_id = task.get('id')
            
            # 检查任务是否与当前输出目录相关
            is_related = False
            args = task.get('args', [])
            kwargs = task.get('kwargs', {})
            
            for arg in args:
                if isinstance(arg, str) and output_dir in arg:
                    is_related = True
                    break
            
            if 'output_dir' in kwargs and output_dir in kwargs['output_dir']:
                is_related = True
            
            if is_related:
                # 获取任务状态
                task_result = AsyncResult(task_id)
                created_time = datetime.fromtimestamp(
                    task.get('time_start', time.time())
                ).strftime('%Y-%m-%d %H:%M:%S')
                
                active_tasks.append({
                    'id': task_id,
                    'state': task_result.state,
                    'created': created_time,
                    'worker': worker_name,
                    'name': '汉字数据导入'
                })
    except Exception as e:
        logger.error(f"获取活动任务列表失败: {str(e)}")
        logger.error(traceback.format_exc())
    
    return active_tasks

@csrf_exempt
def import_data_view(request):
    """处理汉字数据导入请求，支持ZIP图片和两个JSON数据文件（等级和评论）"""
    if request.method == 'POST':
        # 获取上传的文件
        image_zip = request.FILES.get('image_zip')
        json_level = request.FILES.get('json_level')
        json_comment = request.FILES.get('json_comment')
        
        # 获取输出格式
        output_format = request.POST.get('output_format', 'excel')
        
        
        # 打印调试信息
        print(f"收到上传请求 - ZIP: {image_zip.name if image_zip else 'None'}, " +
              f"等级JSON: {json_level.name if json_level else 'None'}, " +
              f"评论JSON: {json_comment.name if json_comment else 'None'}")
        
        # 验证上传的文件
        if not image_zip:
            return JsonResponse({'status': 'error', 'message': '请上传图片ZIP文件'})
            
        # 检查文件类型
        if not image_zip.name.lower().endswith('.zip'):
            return JsonResponse({'status': 'error', 'message': '图片文件必须是ZIP格式'})
            
        # 检查JSON文件格式
        if json_level and not json_level.name.lower().endswith('.json'):
            return JsonResponse({'status': 'error', 'message': '等级JSON文件格式不正确'})
            
        if json_comment and not json_comment.name.lower().endswith('.json'):
            return JsonResponse({'status': 'error', 'message': '评论JSON文件格式不正确'})
        
        # 创建临时目录
        relative_temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_import')
        os.makedirs(relative_temp_dir, exist_ok=True)
        temp_dir = tempfile.mkdtemp(dir=relative_temp_dir, prefix="hanzi_import_")
        print(f"创建临时目录: {temp_dir}")
        
        try:
            # 保存上传的文件
            image_zip_path = os.path.join(temp_dir, image_zip.name)
            with open(image_zip_path, 'wb+') as f:
                for chunk in image_zip.chunks():
                    f.write(chunk)
            print(f"ZIP文件已保存: {image_zip_path}")
            
            # 保存等级JSON文件（如果有）
            json_level_path = None
            if json_level:
                json_level_path = os.path.join(temp_dir, json_level.name)
                with open(json_level_path, 'wb+') as f:
                    for chunk in json_level.chunks():
                        f.write(chunk)
                print(f"等级JSON文件已保存: {json_level_path}")
            
            # 保存评论JSON文件（如果有）
            json_comment_path = None
            if json_comment:
                json_comment_path = os.path.join(temp_dir, json_comment.name)
                with open(json_comment_path, 'wb+') as f:
                    for chunk in json_comment.chunks():
                        f.write(chunk)
                print(f"评论JSON文件已保存: {json_comment_path}")
            
            # 验证文件是否正确保存
            if not os.path.exists(image_zip_path) or os.path.getsize(image_zip_path) == 0:
                return JsonResponse({
                    'status': 'error',
                    'message': f'ZIP文件保存失败或为空: {image_zip_path}'
                })
                
            if json_level_path and (not os.path.exists(json_level_path) or os.path.getsize(json_level_path) == 0):
                return JsonResponse({
                    'status': 'error',
                    'message': f'等级JSON文件保存失败或为空: {json_level_path}'
                })
                
            if json_comment_path and (not os.path.exists(json_comment_path) or os.path.getsize(json_comment_path) == 0):
                return JsonResponse({
                    'status': 'error',
                    'message': f'评论JSON文件保存失败或为空: {json_comment_path}'
                })
                
            # 设置输出目录
            output_dir = os.path.join(settings.MEDIA_ROOT, "import_results")
            os.makedirs(output_dir, exist_ok=True)
            print(f"输出目录: {output_dir}")
            
            # 使用Celery异步任务处理导入
            from .tasks import process_import_data_task
            
            # 启动异步任务并获取任务ID
            task = process_import_data_task.delay(
                image_zip_path,
                json_level_path,
                json_comment_path,
                output_dir,
            )
            
            print(f"Celery任务已提交，任务ID: {task.id}")
            
            # 返回任务ID和状态信息
            return JsonResponse({
                'status': 'processing',
                'message': '数据导入任务已提交，正在后台处理',
                'task_id': task.id
            })
            
        except Exception as e:
            import traceback
            print(f"提交导入任务失败: {str(e)}")
            print(traceback.format_exc())
            
            # 清理临时文件
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)
                
            return JsonResponse({
                'status': 'error',
                'message': f'提交导入任务失败: {str(e)}'
            })
    
    # GET请求返回导入页面
    return render(request, 'hanzi_app/import_data.html', {
        'title': '导入汉字数据'
    })

# 新增任务状态检查API
@require_http_methods(["GET"])
def check_import_task(request):
    """检查导入任务状态"""
    task_id = request.GET.get('task_id')
    
    if not task_id:
        return JsonResponse({'status': 'error', 'message': '缺少任务ID'})
    
    try:
        from .tasks import process_import_data_task
        from celery.result import AsyncResult
        from hanzi_project.celery import app
        import redis
        
        # 获取任务状态
        task_result = AsyncResult(task_id)
        logger.info(f"检查任务 {task_id} 状态：{task_result.state}")
        
        # 从Redis中获取任务日志记录
        task_logs = []
        try:
            redis_client = redis.from_url(app.conf.broker_url)
            task_log_key = f"task_logs:{task_id}"
            logs = redis_client.lrange(task_log_key, 0, -1)
            if logs:
                task_logs = [json.loads(log.decode('utf-8')) for log in logs]
        except Exception as e:
            logger.warning(f"获取任务日志失败: {e}")
        
        # 保存任务状态到session，以便用户退出页面后返回可以继续查看状态
        session_key = f"task_status:{task_id}"
        current_task_status = {
            'id': task_id,
            'state': task_result.state,
            'logs': task_logs,
            'last_check': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        request.session[session_key] = current_task_status
        
        # 从日志中获取最新的进度
        latest_progress = 20  # 默认进度
        latest_message = '导入任务正在处理中'
        
        # 先检查日志记录
        if task_logs:
            # 从最新的日志中提取进度信息
            for log in reversed(task_logs):
                if 'progress' in log:
                    latest_progress = log['progress']
                    if 'message' in log:
                        latest_message = log['message']
                    break
        
        # 根据任务状态判断进度
        if not task_logs and task_result.state == 'PENDING':
            # 如果没有日志且状态是PENDING，设置初始进度
            latest_progress = 5
            latest_message = '任务排队中...'
        elif task_result.state == 'STARTED':
            # 根据状态设置最小进度
            if latest_progress < 10:
                latest_progress = 10
                latest_message = '任务已开始，正在解压文件...'
        
        # 如果是已知的进度百分比字符串，从中提取实际数字
        if isinstance(latest_message, str) and '进度' in latest_message and '%' in latest_message:
            try:
                # 尝试从消息中提取进度数字
                import re
                progress_match = re.search(r'(\d+)%', latest_message)
                if progress_match:
                    extracted_progress = int(progress_match.group(1))
                    if extracted_progress > latest_progress:
                        latest_progress = extracted_progress
            except Exception:
                pass
        
        # 构建响应结果
        result = {
            'task_id': task_id,
            'state': task_result.state,
            'status': 'processing',
            'message': latest_message,
            'progress': latest_progress,
            'task_logs': task_logs[-5:] if task_logs else []  # 只返回最近的5条日志
        }
        
        # 更新进度状态
        if task_result.state == 'PENDING':
            if latest_progress < 10:
                result['progress'] = 5
            result['message'] = latest_message or '任务排队中...'
        elif task_result.state == 'STARTED':
            # 确保进度至少为20%
            result['progress'] = max(latest_progress, 20)
            result['message'] = latest_message or '任务处理中...'
        elif task_result.state == 'RETRY':
            result['progress'] = max(latest_progress, 40)
            result['message'] = '任务正在重试...'
        
        # 处理已完成的任务
        if task_result.ready():
            if task_result.successful():
                # 任务成功完成，获取结果
                try:
                    task_data = task_result.result
                    
                    # 检查结果是否为字典
                    if isinstance(task_data, dict):
                        result['status'] = task_data.get('status', 'completed')
                        result['message'] = task_data.get('message', '数据导入完成')
                        result['file_url'] = task_data.get('file_url')
                        result['processed_count'] = task_data.get('processed_count', 0)
                        result['recognized_count'] = task_data.get('recognized_count', 0)
                        result['progress'] = 100
                        
                        # 检查结果文件是否存在
                        if result.get('file_url'):
                            media_url = '/media/'
                            if result['file_url'].startswith(media_url):
                                url_path = result['file_url'][len(media_url):]
                                file_path = os.path.normpath(os.path.join(settings.MEDIA_ROOT, url_path))
                                
                                if os.path.exists(file_path):
                                    result['file_exists'] = True
                                    result['file_size'] = os.path.getsize(file_path)
                                else:
                                    result['file_exists'] = False
                                    logger.warning(f"结果文件不存在：{file_path}")
                            else:
                                result['file_exists'] = False
                                logger.warning(f"无效的文件URL：{result['file_url']}")
                except Exception as e:
                    logger.error(f"获取任务结果时出错: {str(e)}")
                    result['status'] = 'error'
                    result['message'] = f'获取任务结果时出错: {str(e)}'
            else:
                # 任务失败
                result['status'] = 'failed'
                result['message'] = '任务执行失败'
                try:
                    if task_result.result:
                        result['error'] = str(task_result.result)
                except Exception:
                    pass
        
        # 获取import_results目录中所有可下载的文件
        import_results_dir = os.path.join(settings.MEDIA_ROOT, "import_results")
        completed_files = get_completed_files(import_results_dir)
        if completed_files:
            result['completed_files'] = completed_files[:10]  # 只返回最近的10个文件
        
        return JsonResponse(result)
    except Exception as e:
        logger.error(f"检查任务状态时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({
            'status': 'error',
            'message': f'检查任务状态时出错: {str(e)}',
            'task_id': task_id
        })

@require_http_methods(["GET"])
def check_import_status(request):
    """检查导入状态，返回导入任务列表和可用的导入结果文件"""
    import_results_dir = os.path.join(settings.MEDIA_ROOT, "import_results")
    
    try:
        # 获取可用的导入结果文件
        files = get_completed_files(import_results_dir)
        
        # 获取活动任务
        tasks = []
        for key, value in request.session.items():
            if key.startswith('task_status:'):
                task_id = key.split(':', 1)[1]
                # 检查是否为最近24小时内的任务
                last_check = value.get('last_check')
                if last_check:
                    try:
                        last_check_time = datetime.strptime(last_check, '%Y-%m-%d %H:%M:%S')
                        time_diff = datetime.now() - last_check_time
                        if time_diff.total_seconds() < 86400:  # 24小时内
                            tasks.append({
                                'id': task_id,
                                'state': value.get('state'),
                                'last_check': last_check
                            })
                    except Exception:
                        pass
        
        # 检查临时目录
        temp_import_dir = os.path.join(settings.MEDIA_ROOT, "temp_import")
        import_folders = []
        expired_folders = []
        
        if os.path.exists(temp_import_dir):
            try:
                for item in os.listdir(temp_import_dir):
                    full_path = os.path.join(temp_import_dir, item)
                    if os.path.isdir(full_path) and (item.startswith('hanzi_import_') or item.startswith('import_task_') or item.startswith('tmp')):
                        # 获取目录信息
                        mod_time = datetime.fromtimestamp(os.path.getmtime(full_path))
                        time_diff = datetime.now() - mod_time
                        
                        # 计算时间差（小时）
                        age_hours = round(time_diff.total_seconds() / 3600, 1)
                        
                        # 超过48小时的文件夹视为过期
                        if age_hours > 48:
                            expired_folders.append(full_path)
                        elif time_diff.total_seconds() < 86400 * 7:  # 7天内的非过期文件夹
                            import_folders.append({
                                'name': item,
                                'path': full_path,
                                'created': mod_time.strftime('%Y-%m-%d %H:%M:%S'),
                                'age_hours': age_hours
                            })
            except Exception as e:
                logger.warning(f"获取导入临时文件夹失败: {e}")
                
        # 清理过期的临时文件夹（后台操作）
        if expired_folders:
            try:
                for folder in expired_folders:
                    try:
                        shutil.rmtree(folder, ignore_errors=True)
                        logger.info(f"已清理过期的临时文件夹: {folder}")
                    except Exception as e:
                        logger.warning(f"清理过期文件夹失败: {folder}, 错误: {e}")
            except Exception as e:
                logger.warning(f"清理过期文件夹过程中出错: {e}")
        
        return JsonResponse({
            'status': 'success',
            'files': files,
            'tasks': tasks,
            'import_folders': import_folders[:10],  # 只返回最近的10个文件夹
            'cleaned_folders': len(expired_folders)
        })
    except Exception as e:
        logger.error(f"检查导入状态失败: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': f'检查导入状态失败: {str(e)}'
        })

@require_http_methods(["POST"])
@csrf_exempt
def delete_import_file(request):
    """删除导入结果文件"""
    filename = request.POST.get('filename')
    
    if not filename:
        return JsonResponse({'success': False, 'error': '未提供文件名'})
    
    # 安全检查 - 确保文件名符合预期格式
    if not filename.startswith('hanzi_import_') or not filename.endswith('.xlsx'):
        return JsonResponse({'success': False, 'error': '无效的文件名'})
    
    # 构建文件路径
    file_path = os.path.join(settings.MEDIA_ROOT, "import_results", filename)
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        return JsonResponse({'success': False, 'error': '文件不存在'})
    
    try:
        # 删除Excel文件
        os.remove(file_path)
        
        # 同时删除相关的日志文件（如果存在）
        log_filename = os.path.splitext(filename)[0] + '_failed.log'
        log_path = os.path.join(settings.MEDIA_ROOT, "import_results", log_filename)
        if os.path.exists(log_path):
            os.remove(log_path)
        
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"删除文件失败: {str(e)}")
        return JsonResponse({'success': False, 'error': f'删除文件失败: {str(e)}'})